from bs4 import BeautifulSoup
from urllib.request import urlopen
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import xlrd
from openpyxl import load_workbook



login = ''
password = ''

def anbient(link):
    print('Capturando links dos episódios...')

    #driver = webdriver.Chrome()
   # driver.get('https://www.anbient.com/user/login')
   # username = driver.find_element_by_xpath('//*[@id="edit-name"]')
   # username.send_keys(login)
   # username = driver.find_element_by_xpath('//*[@id="edit-pass"]')
   # username.send_keys(password)
   # time.sleep(1)
   # driver.find_element(By.XPATH, '//*[@id="edit-submit"]').click()

    driver.get(link)

    ids = driver.page_source

    soup = BeautifulSoup(ids, 'html.parser')

    busc = soup.find_all("li")
    txt = str(busc).split('"')
    #print(txt)
    contar_lista_txt = len(txt)
    #print(contar_lista_txt)
    lista_links = []
    for c in range(0, contar_lista_txt):
        #print(lista_links)
        fim = txt[c].find('anbient.com')
        if fim != -1:
            lista_links.append(txt[c])
    print(lista_links)
    del(lista_links[4])
    del(lista_links[3])
    del(lista_links[2])
    del(lista_links[1])
    del(lista_links[0])
    print(lista_links)        

    if len(lista_links) == 0:
        print()
        print('Provavelmente o serviço de download zippyshare não está disponível')
        retorno()
    for empilhados in range(0, len(lista_links)):
        link_escolhido = f'{lista_links[empilhados]}'
        print(link_escolhido)
        time.sleep(2)
        driver.get(link_escolhido)
        id = driver.page_source
        sopa = BeautifulSoup(id, 'html.parser')
        #zip_link = sopa.find_all("a", id=True)
        picotado = str(link_escolhido).split('/')
        print(picotado)
        driver.get(f'https://{picotado[2]}{zip}')
        

def zippyshare(link):
    #driver = webdriver.Chrome()
    print('Capturando links dos episódios...')
    driver.get(link)
    ids = driver.page_source
    soup = BeautifulSoup(ids, 'html.parser')
    busc = soup.find_all("li")
    txt = str(busc).split('"')
    contar_lista_txt = len(txt)
    lista_links = []
    for c in range(0, contar_lista_txt):
        fim = txt[c].find('zippyshare.com')
        if fim != -1:
            lista_links.append(txt[c])        
    if len(lista_links) == 0:
        print()
        print('Provavelmente o serviço de download zippyshare não está disponível')
        retorno()


    for empilhados in range(0, len(lista_links)):
        link_escolhido = f'{lista_links[empilhados]}'
        print(link_escolhido)
        driver.get(link_escolhido)
        id = driver.page_source
        sopa = BeautifulSoup(id, 'html.parser')
        zip_link = sopa.find_all("a", id=True)
        zip = zip_link[0].get('href')
        
    
        picotado = str(link_escolhido).split('/')
        driver.get(f'https://{picotado[2]}{zip}')


#Abrir excel
wb = load_workbook('ListaAnime.xlsx')
ws = wb['Erros']

#List Anime excel
book = xlrd.open_workbook("ListaAnime.xlsx")
sh = book.sheet_by_index(0)

driver = webdriver.Chrome()
driver.get('https://www.anbient.com/user/login')
username = driver.find_element_by_xpath('//*[@id="edit-name"]')
username.send_keys(login)
username = driver.find_element_by_xpath('//*[@id="edit-pass"]')
username.send_keys(password)
time.sleep(1)
driver.find_element(By.XPATH, '//*[@id="edit-submit"]').click()

for rx in range(sh.nrows):
    print("____________________________________________________________")
    print("Name: ", sh.cell_value(rowx=rx, colx=0))
    print("EP: ", sh.cell_value(rowx=rx, colx=3))
    print("LINK: ", sh.cell_value(rowx=rx, colx=4))
    print("____________________________________________________________")
    print()
    print()
    try:
        zippyshare(sh.cell_value(rowx=rx, colx=4))
    except:
        try:
            anbient(sh.cell_value(rowx=rx, colx=4))
        except:
            ws['A'+str(rx+1)] = sh.cell_value(rowx=rx, colx=0)
            ws['D'+str(rx+1)] = sh.cell_value(rowx=rx, colx=3)
            ws['E'+str(rx+1)] = sh.cell_value(rowx=rx, colx=4)
            wb.save('ListaAnimeErro.xlsx')    
